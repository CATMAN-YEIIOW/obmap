from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, and_
from typing import Optional, List
from datetime import datetime, timedelta, timezone
from pydantic import BaseModel
from zoneinfo import ZoneInfo
import math
import csv
import io

BEIJING_TZ = ZoneInfo("Asia/Shanghai")

# 采集间隔（秒）
SAMPLING_INTERVAL_SECONDS = 10

# 报表类型
class ReportType:
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"


def to_beijing_time(dt: datetime) -> datetime:
    """将UTC datetime转换为北京时间"""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(BEIJING_TZ)


def format_beijing_time(dt: datetime) -> str:
    """格式化为北京时间字符串"""
    return to_beijing_time(dt).strftime("%Y-%m-%d %H:%M:%S")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

from app.database import get_db
from app.models import Buoy, BuoyData, BuoyStatusLog, BuoyStatus

router = APIRouter(prefix="/api/v1/statistics", tags=["统计分析"])


class StatValue(BaseModel):
    min: float
    max: float
    avg: float
    std: Optional[float] = None
    unit: str


class StatisticsSummary(BaseModel):
    buoy_id: str
    period: dict
    records: int
    statistics: dict


PARAM_CONFIG = {
    "temperature": {"unit": "°C"},
    "salinity": {"unit": "PSU"},
    "ph": {"unit": ""},
    "dissolved_oxygen": {"unit": "mg/L"},
    "turbidity": {"unit": "NTU"},
    "chlorophyll": {"unit": "μg/L"},
    "wave_height": {"unit": "m"},
}


@router.get("/summary")
def get_statistics_summary(
    buoy_id: str = Query(..., description="浮标ID"),
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    db: Session = Depends(get_db)
):
    buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
    if not buoy:
        raise HTTPException(status_code=404, detail="浮标不存在")

    records = db.query(BuoyData).filter(
        BuoyData.buoy_id == buoy_id,
        BuoyData.time >= start_time,
        BuoyData.time <= end_time,
        BuoyData.drift_flagged == False
    ).count()

    statistics = {}
    for param, config in PARAM_CONFIG.items():
        column = getattr(BuoyData, param)
        stats = db.query(
            func.min(column).label("min"),
            func.max(column).label("max"),
            func.avg(column).label("avg"),
            func.stddev(column).label("std")
        ).filter(
            BuoyData.buoy_id == buoy_id,
            BuoyData.time >= start_time,
            BuoyData.time <= end_time,
            column.isnot(None),
            BuoyData.drift_flagged == False
        ).first()

        if stats and stats.min is not None:
            statistics[param] = {
                "min": float(stats.min),
                "max": float(stats.max),
                "avg": round(float(stats.avg), 2),
                "std": round(float(stats.std), 2) if stats.std else None,
                "unit": config["unit"]
            }

    return {
        "code": 200,
        "message": "success",
        "data": {
            "buoy_id": str(buoy_id),
            "buoy_name": buoy.name,
            "period": {
                "start": start_time.isoformat(),
                "end": end_time.isoformat()
            },
            "records": records,
            "statistics": statistics
        }
    }


@router.get("/timeseries")
def get_timeseries_statistics(
    buoy_id: str = Query(..., description="浮标ID"),
    param: str = Query(..., description="参数名"),
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    bucket: str = Query("1h", description="聚合间隔: 1h/6h/1d"),
    db: Session = Depends(get_db)
):
    buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
    if not buoy:
        raise HTTPException(status_code=404, detail="浮标不存在")

    interval_map = {
        "1h": "1 hour",
        "6h": "6 hour",
        "1d": "1 day",
        "30m": "30 min",
        "12h": "12 hour"
    }
    interval = interval_map.get(bucket, "1 hour")

    if not hasattr(BuoyData, param):
        raise HTTPException(status_code=400, detail=f"参数 {param} 不存在")

    # Use TimescaleDB's time_bucket for arbitrary intervals
    bucket_time = func.time_bucket(interval, BuoyData.time).label('bucket_time')

    query = db.query(
        bucket_time,
        func.min(getattr(BuoyData, param)).label('min_val'),
        func.max(getattr(BuoyData, param)).label('max_val'),
        func.avg(getattr(BuoyData, param)).label('avg_val'),
        func.count(getattr(BuoyData, param)).label('count')
    ).filter(
        BuoyData.buoy_id == buoy_id,
        BuoyData.time >= start_time,
        BuoyData.time <= end_time,
        getattr(BuoyData, param).isnot(None),
        BuoyData.drift_flagged == False
    ).group_by(bucket_time).order_by(bucket_time)

    result = query.all()

    return {
        "code": 200,
        "message": "success",
        "data": {
            "param": param,
            "unit": PARAM_CONFIG.get(param, {}).get("unit", ""),
            "bucket": bucket,
            "buoy_name": buoy.name,
            "items": [
                {
                    "time": row.bucket_time.isoformat() if row.bucket_time else str(row[0]),
                    "min": float(row.min_val) if row.min_val is not None else None,
                    "max": float(row.max_val) if row.max_val is not None else None,
                    "avg": round(float(row.avg_val), 2) if row.avg_val is not None else None,
                    "count": row.count
                } for row in result
            ]
        }
    }


@router.get("/raw")
def get_raw_data(
    buoy_id: str = Query(..., description="浮标ID"),
    param: str = Query(..., description="参数名"),
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    db: Session = Depends(get_db)
):
    """获取原始数据（无聚合），适用于短周期视图"""
    buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
    if not buoy:
        raise HTTPException(status_code=404, detail="浮标不存在")

    if not hasattr(BuoyData, param):
        raise HTTPException(status_code=400, detail=f"参数 {param} 不存在")

    column = getattr(BuoyData, param)

    # 直接查询原始数据，按时间正序
    query = db.query(
        BuoyData.time,
        column
    ).filter(
        BuoyData.buoy_id == buoy_id,
        BuoyData.time >= start_time,
        BuoyData.time <= end_time,
        column.isnot(None),
        BuoyData.drift_flagged == False
    ).order_by(BuoyData.time)

    result = query.all()

    return {
        "code": 200,
        "message": "success",
        "data": {
            "param": param,
            "unit": PARAM_CONFIG.get(param, {}).get("unit", ""),
            "buoy_name": buoy.name,
            "items": [
                {
                    "time": row.time.isoformat(),
                    "value": float(row[1]) if row[1] is not None else None
                } for row in result
            ]
        }
    }


@router.get("/thresholds")
def get_buoy_thresholds(
    buoy_id: str = Query(..., description="浮标ID"),
    db: Session = Depends(get_db)
):
    """获取浮标的告警阈值配置（包括全局配置和浮标专属配置）"""
    from app.models import AlertConfig, AlertSeverity

    # 优先获取浮标专属配置，如果没有则使用全局配置
    configs = db.query(AlertConfig).filter(
        AlertConfig.enabled == True,
        (AlertConfig.buoy_id == buoy_id) | (AlertConfig.buoy_id == None)
    ).all()

    # 按优先级排序：浮标专属配置优先于全局配置
    # 先按 buoy_id DESC 排序，让专属配置（在前面）覆盖全局配置
    configs_sorted = sorted(configs, key=lambda x: (x.buoy_id is not None, x.buoy_id or ''), reverse=True)

    thresholds = {}
    for cfg in configs_sorted:
        if cfg.param_name:
            # 如果已经存在阈值（来自全局配置），跳过（因为专属配置已处理过）
            if cfg.param_name not in thresholds:
                thresholds[cfg.param_name] = {
                    "min_threshold": float(cfg.min_threshold) if cfg.min_threshold else None,
                    "max_threshold": float(cfg.max_threshold) if cfg.max_threshold else None,
                    "severity": cfg.severity.value if cfg.severity else "warning"
                }

    return {
        "code": 200,
        "message": "success",
        "data": {
            "buoy_id": str(buoy_id),
            "thresholds": thresholds
        }
    }


@router.get("/alert-events")
def get_alert_events(
    buoy_id: str = Query(..., description="浮标ID"),
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    db: Session = Depends(get_db)
):
    """获取时间范围内的告警事件，用于在图表上标注"""
    from app.models import Alert

    alerts = db.query(Alert).filter(
        Alert.buoy_id == buoy_id,
        Alert.triggered_at >= start_time,
        Alert.triggered_at <= end_time
    ).order_by(Alert.triggered_at).all()

    return {
        "code": 200,
        "message": "success",
        "data": {
            "buoy_id": str(buoy_id),
            "items": [
                {
                    "id": str(a.id),
                    "param_name": a.param_name,
                    "alert_type": a.alert_type,
                    "severity": a.severity.value if a.severity else "warning",
                    "actual_value": float(a.actual_value) if a.actual_value else None,
                    "threshold_value": float(a.threshold_value) if a.threshold_value else None,
                    "triggered_at": a.triggered_at.isoformat() if a.triggered_at else None
                } for a in alerts
            ]
        }
    }


@router.get("/compare")
def compare_buoys(
    buoy_ids: str = Query(..., description="逗号分隔的浮标ID列表"),
    param: str = Query(..., description="参数名"),
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    db: Session = Depends(get_db)
):
    id_list = [id.strip() for id in buoy_ids.split(",")]
    if len(id_list) > 5:
        raise HTTPException(status_code=400, detail="最多支持5个浮标对比")

    column = getattr(BuoyData, param)
    result = []

    for buoy_id in id_list:
        buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
        if not buoy:
            continue

        stats = db.query(
            func.min(column).label("min"),
            func.max(column).label("max"),
            func.avg(column).label("avg")
        ).filter(
            BuoyData.buoy_id == buoy_id,
            BuoyData.time >= start_time,
            BuoyData.time <= end_time,
            column.isnot(None),
            BuoyData.drift_flagged == False
        ).first()

        if stats and stats.min is not None:
            result.append({
                "buoy_id": str(buoy_id),
                "buoy_name": buoy.name,
                "sea_area": buoy.sea_area,
                "avg": round(float(stats.avg), 2),
                "min": float(stats.min),
                "max": float(stats.max),
                "unit": PARAM_CONFIG.get(param, {}).get("unit", "")
            })

    return {
        "code": 200,
        "message": "success",
        "data": {
            "param": param,
            "buoys": result
        }
    }


@router.get("/period")
def get_multi_period_statistics(
    buoy_id: str = Query(..., description="浮标ID"),
    period: str = Query("day", description="统计周期: day/week/month"),
    db: Session = Depends(get_db)
):
    """多周期统计 - 按日/周/月聚合统计"""
    buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
    if not buoy:
        raise HTTPException(status_code=404, detail="浮标不存在")

    now = datetime.utcnow()

    if period == "day":
        start_time = now - timedelta(days=7)
        bucket = "hour"
        period_label = "近7天（小时级）"
    elif period == "week":
        start_time = now - timedelta(weeks=4)
        bucket = "day"
        period_label = "近4周（日级）"
    elif period == "month":
        start_time = now - timedelta(days=90)
        bucket = "day"
        period_label = "近3个月（日级）"
    else:
        raise HTTPException(status_code=400, detail="无效的周期")

    interval_map = {"hour": "hour", "day": "day"}
    bucket_time = func.date_trunc(interval_map.get(bucket, "day"), BuoyData.time).label('bucket_time')

    all_data = []
    for param in PARAM_CONFIG.keys():
        query = db.query(
            bucket_time,
            func.min(getattr(BuoyData, param)).label('min_val'),
            func.max(getattr(BuoyData, param)).label('max_val'),
            func.avg(getattr(BuoyData, param)).label('avg_val'),
            func.count(getattr(BuoyData, param)).label('count')
        ).filter(
            BuoyData.buoy_id == buoy_id,
            BuoyData.time >= start_time,
            BuoyData.time <= now,
            getattr(BuoyData, param).isnot(None),
            BuoyData.drift_flagged == False
        ).group_by(bucket_time).order_by(bucket_time)

        result = query.all()
        all_data.append({
            "param": param,
            "unit": PARAM_CONFIG[param]["unit"],
            "items": [
                {
                    "time": row.bucket_time.isoformat() if row.bucket_time else str(row[0]),
                    "min": float(row.min_val) if row.min_val is not None else None,
                    "max": float(row.max_val) if row.max_val is not None else None,
                    "avg": round(float(row.avg_val), 2) if row.avg_val is not None else None,
                    "count": row.count
                } for row in result
            ]
        })

    return {
        "code": 200,
        "message": "success",
        "data": {
            "buoy_id": str(buoy_id),
            "buoy_name": buoy.name,
            "period_type": period,
            "period_label": period_label,
            "start_time": start_time.isoformat(),
            "end_time": now.isoformat(),
            "params": all_data
        }
    }


@router.get("/export/data")
def export_data(
    buoy_id: str = Query(..., description="浮标ID"),
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    format: str = Query("csv", description="导出格式: csv/xlsx"),
    db: Session = Depends(get_db)
):
    buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
    if not buoy:
        raise HTTPException(status_code=404, detail="浮标不存在")

    # 导出包含所有数据（包括特殊状态的空数据行），用于显示完整的时间序列
    data = db.query(BuoyData).filter(
        BuoyData.buoy_id == buoy_id,
        BuoyData.time >= start_time,
        BuoyData.time <= end_time
    ).order_by(BuoyData.time).all()

    if format == "xlsx" and HAS_XLSX:
        return _export_excel(buoy, data, start_time, end_time)
    else:
        return _export_csv(buoy, data, start_time, end_time)


def _export_csv(buoy: Buoy, data, start_time, end_time):
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "时间", "水温(°C)", "盐度(PSU)", "pH值",
        "溶解氧(mg/L)", "浊度(NTU)", "叶绿素(μg/L)", "波高(m)"
    ])

    for row in data:
        writer.writerow([
            format_beijing_time(row.time),
            row.temperature,
            row.salinity,
            row.ph,
            row.dissolved_oxygen,
            row.turbidity,
            row.chlorophyll,
            row.wave_height
        ])

    output.seek(0)
    filename = f"buoy_data_{buoy.code}_{to_beijing_time(start_time).strftime('%Y%m%d')}_{to_beijing_time(end_time).strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _export_excel(buoy: Buoy, data, start_time, end_time):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{buoy.code}监测数据"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    headers = ["时间", "水温(°C)", "盐度(PSU)", "pH值", "溶解氧(mg/L)", "浊度(NTU)", "叶绿素(μg/L)", "波高(m)"]
    ws.append(headers)

    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in data:
        ws.append([
            format_beijing_time(row.time),
            row.temperature,
            row.salinity,
            row.ph,
            row.dissolved_oxygen,
            row.turbidity,
            row.chlorophyll,
            row.wave_height
        ])

    # Auto column width
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max_length + 2, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"buoy_data_{buoy.code}_{to_beijing_time(start_time).strftime('%Y%m%d')}_{to_beijing_time(end_time).strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """计算两点之间的Haversine距离（公里）"""
    R = 6371  # 地球半径（公里）
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + \
        math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * \
        math.sin(dlon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


@router.get("/status")
def get_buoy_status_statistics(
    buoy_id: str = Query(..., description="浮标ID"),
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    db: Session = Depends(get_db)
):
    """获取浮标状态统计，包括状态时长、变更记录、漂移事件、数据完整率"""
    buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
    if not buoy:
        raise HTTPException(status_code=404, detail="浮标不存在")

    # 1. 获取状态变更日志
    status_logs = db.query(BuoyStatusLog).filter(
        BuoyStatusLog.buoy_id == buoy_id,
        BuoyStatusLog.changed_at >= start_time,
        BuoyStatusLog.changed_at <= end_time
    ).order_by(BuoyStatusLog.changed_at).all()

    # 2. 计算各状态时长（小时）
    status_hours = {
        "online": 0.0,
        "offline": 0.0,
        "disconnected": 0.0,
        "low_battery": 0.0,
        "no_power": 0.0,
        "drift_alert": 0.0,
        "inactive": 0.0
    }

    total_seconds = (end_time - start_time).total_seconds()
    total_hours = total_seconds / 3600

    if not status_logs:
        # 没有状态变更记录，使用当前状态估算
        current_status = buoy.status.value if hasattr(buoy.status, 'value') else buoy.status
        if current_status in status_hours:
            status_hours[current_status] = total_hours
    else:
        # 计算每段状态的时长
        prev_log = None
        for log in status_logs:
            if prev_log is None:
                # 第一条记录：从开始时间到第一次状态变更
                if log.changed_at > start_time:
                    duration = (log.changed_at - start_time).total_seconds() / 3600
                    prev_status = log.previous_status.value if hasattr(log.previous_status, 'value') else log.previous_status
                    if prev_status in status_hours:
                        status_hours[prev_status] += duration
            else:
                # 上一条记录的状态到当前记录的状态之间的时间
                duration = (log.changed_at - prev_log.changed_at).total_seconds() / 3600
                prev_status = log.previous_status.value if hasattr(log.previous_status, 'value') else log.previous_status
                if prev_status in status_hours:
                    status_hours[prev_status] += duration

            prev_log = log

        # 最后一段：从最后一条记录到结束时间（假设保持最后的状态）
        last_status = status_logs[-1].status.value if hasattr(status_logs[-1].status, 'value') else status_logs[-1].status
        if status_logs[-1].changed_at < end_time:
            duration = (end_time - status_logs[-1].changed_at).total_seconds() / 3600
            if last_status in status_hours:
                status_hours[last_status] += duration

    # 3. 构建状态变更记录
    status_changes = []
    for log in status_logs:
        prev_status = log.previous_status.value if hasattr(log.previous_status, 'value') else log.previous_status
        curr_status = log.status.value if hasattr(log.status, 'value') else log.status
        status_changes.append({
            "changed_at": log.changed_at.isoformat() if log.changed_at else None,
            "previous_status": prev_status,
            "new_status": curr_status,
            "reason": log.reason,
            "latitude": float(log.latitude) if log.latitude else None,
            "longitude": float(log.longitude) if log.longitude else None,
            "battery_level": log.battery_level
        })

    # 4. 计算漂移事件（连续漂移状态为一组）
    drift_events = []
    current_drift = None

    for log in status_logs:
        curr_status = log.status.value if hasattr(log.status, 'value') else log.status
        if curr_status == "drift_alert":
            if current_drift is None:
                current_drift = {
                    "start": log.changed_at.isoformat() if log.changed_at else None,
                    "start_lat": float(log.latitude) if log.latitude else None,
                    "start_lon": float(log.longitude) if log.longitude else None,
                    "start_battery": log.battery_level,
                    "max_offset_km": 0.0,
                    "max_offset_lat": float(log.latitude) if log.latitude else None,
                    "max_offset_lon": float(log.longitude) if log.longitude else None
                }
            else:
                # 更新最大偏移
                if log.latitude and log.longitude and current_drift["start_lat"] and current_drift["start_lon"]:
                    offset = haversine_distance(
                        current_drift["start_lat"], current_drift["start_lon"],
                        float(log.latitude), float(log.longitude)
                    )
                    if offset > current_drift["max_offset_km"]:
                        current_drift["max_offset_km"] = round(offset, 4)
                        current_drift["max_offset_lat"] = float(log.latitude)
                        current_drift["max_offset_lon"] = float(log.longitude)
        else:
            if current_drift is not None:
                current_drift["end"] = log.changed_at.isoformat() if log.changed_at else None
                current_drift["end_lat"] = float(log.latitude) if log.latitude else None
                current_drift["end_lon"] = float(log.longitude) if log.longitude else None
                current_drift["max_offset_km"] = round(current_drift["max_offset_km"], 4)
                drift_events.append(current_drift)
                current_drift = None

    # 如果漂移状态持续到结束时间
    if current_drift is not None:
        current_drift["end"] = end_time.isoformat()
        current_drift["end_lat"] = float(buoy.latitude) if buoy.latitude else None
        current_drift["end_lon"] = float(buoy.longitude) if buoy.longitude else None
        current_drift["max_offset_km"] = round(current_drift["max_offset_km"], 4)
        drift_events.append(current_drift)

    # 5. 计算数据完整率
    expected_points = int(total_seconds / SAMPLING_INTERVAL_SECONDS)
    actual_points = db.query(BuoyData).filter(
        BuoyData.buoy_id == buoy_id,
        BuoyData.time >= start_time,
        BuoyData.time <= end_time
    ).count()

    completeness_rate = round((actual_points / expected_points * 100), 2) if expected_points > 0 else 0

    # 6. 统计状态变更次数
    status_change_count = len(status_logs)

    # 7. 计算最长连续在线/离线时间
    longest_continuous = {"online": 0.0, "offline": 0.0, "disconnected": 0.0}
    current_streak = {"status": None, "start": None, "duration": 0.0}

    all_logs_sorted = sorted(status_logs, key=lambda x: x.changed_at)

    # 加入开始时间作为起点
    if all_logs_sorted:
        first_log = all_logs_sorted[0]
        if first_log.changed_at > start_time:
            prev_status = first_log.previous_status.value if hasattr(first_log.previous_status, 'value') else first_log.previous_status
            current_streak = {"status": prev_status, "start": start_time, "duration": (first_log.changed_at - start_time).total_seconds() / 3600}

    for log in all_logs_sorted:
        curr_status = log.status.value if hasattr(log.status, 'value') else log.status

        if current_streak["status"] == curr_status:
            # 同一状态持续
            current_streak["duration"] += (log.changed_at - (log.changed_at - timedelta(hours=current_streak["duration"]))).total_seconds() / 3600
        else:
            # 状态变更，检查是否是 online/offline/disconnected
            if current_streak["status"] in longest_continuous:
                if current_streak["duration"] > longest_continuous[current_streak["status"]]:
                    longest_continuous[current_streak["status"]] = round(current_streak["duration"], 2)
            # 开始新的连续状态
            current_streak = {"status": curr_status, "start": log.changed_at, "duration": 0.0}

    # 检查最后一段
    if current_streak["status"] in longest_continuous and status_logs:
        last_log = status_logs[-1]
        if last_log.changed_at < end_time:
            duration = (end_time - last_log.changed_at).total_seconds() / 3600
            current_streak["duration"] += duration
        if current_streak["duration"] > longest_continuous[current_streak["status"]]:
            longest_continuous[current_streak["status"]] = round(current_streak["duration"], 2)

    return {
        "code": 200,
        "message": "success",
        "data": {
            "buoy_id": str(buoy_id),
            "buoy_name": buoy.name,
            "period": {
                "start": start_time.isoformat(),
                "end": end_time.isoformat(),
                "total_hours": round(total_hours, 2)
            },
            "status_summary": {
                "online_hours": round(status_hours.get("online", 0), 2),
                "offline_hours": round(status_hours.get("offline", 0), 2),
                "disconnected_hours": round(status_hours.get("disconnected", 0), 2),
                "low_battery_hours": round(status_hours.get("low_battery", 0), 2),
                "no_power_hours": round(status_hours.get("no_power", 0), 2),
                "drift_alert_hours": round(status_hours.get("drift_alert", 0), 2),
                "inactive_hours": round(status_hours.get("inactive", 0), 2),
                "online_rate": round((status_hours.get("online", 0) / total_hours * 100), 2) if total_hours > 0 else 0
            },
            "status_change_count": status_change_count,
            "longest_continuous": {
                "online_hours": longest_continuous.get("online", 0),
                "offline_hours": longest_continuous.get("offline", 0),
                "disconnected_hours": longest_continuous.get("disconnected", 0)
            },
            "status_changes": status_changes,
            "drift_events": drift_events,
            "data_completeness": {
                "sampling_interval_seconds": SAMPLING_INTERVAL_SECONDS,
                "expected_points": expected_points,
                "actual_points": actual_points,
                "completeness_rate": completeness_rate,
                "missing_points": expected_points - actual_points
            }
        }
    }


# ===== 报表生成 =====
class ReportGenerateRequest(BaseModel):
    buoy_ids: List[str] = Query(..., description="浮标ID列表")
    report_type: str = Query(..., description="报表类型: daily/weekly/monthly/quarterly")
    start_time: Optional[datetime] = Query(None, description="自定义开始时间（可选，不填则根据报表类型自动计算）")
    end_time: Optional[datetime] = Query(None, description="自定义结束时间（可选，不填则根据报表类型自动计算）")
    include_trends: bool = Query(True, description="是否包含趋势图")


def get_report_time_range(report_type: str, reference_time: datetime = None):
    """根据报表类型计算时间范围"""
    if reference_time is None:
        reference_time = datetime.utcnow()

    if report_type == ReportType.DAILY:
        start = reference_time.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1) - timedelta(seconds=1)
    elif report_type == ReportType.WEEKLY:
        days_since_monday = reference_time.weekday()
        start = (reference_time - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7) - timedelta(seconds=1)
    elif report_type == ReportType.MONTHLY:
        start = reference_time.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if reference_time.month == 12:
            end = reference_time.replace(year=reference_time.year + 1, month=1, day=1) - timedelta(seconds=1)
        else:
            end = reference_time.replace(month=reference_time.month + 1, day=1) - timedelta(seconds=1)
    elif report_type == ReportType.QUARTERLY:
        quarter = (reference_time.month - 1) // 3
        start_month = quarter * 3 + 1
        start = reference_time.replace(month=start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        if start_month == 10:
            end = reference_time.replace(year=reference_time.year + 1, month=1, day=1) - timedelta(seconds=1)
        else:
            end = reference_time.replace(month=start_month + 3, day=1) - timedelta(seconds=1)
    else:
        raise HTTPException(status_code=400, detail=f"无效的报表类型: {report_type}")

    return start, end


@router.post("/report")
def generate_report(
    request: ReportGenerateRequest,
    db: Session = Depends(get_db)
):
    """生成 PDF 统计报表"""
    # 验证报表类型
    if request.report_type not in [ReportType.DAILY, ReportType.WEEKLY, ReportType.MONTHLY, ReportType.QUARTERLY]:
        raise HTTPException(status_code=400, detail="无效的报表类型，支持: daily/weekly/monthly/quarterly")

    # 验证浮标
    if not request.buoy_ids:
        raise HTTPException(status_code=400, detail="请至少选择一个浮标")

    for buoy_id in request.buoy_ids:
        buoy = db.query(Buoy).filter(Buoy.id == buoy_id).first()
        if not buoy:
            raise HTTPException(status_code=404, detail=f"浮标 {buoy_id} 不存在")

    # 计算时间范围
    if request.start_time and request.end_time:
        start_time = request.start_time
        end_time = request.end_time
    else:
        start_time, end_time = get_report_time_range(request.report_type)

    # 导入报表生成服务
    from app.services.report import ReportGenerator, ReportConfig

    config = ReportConfig(
        buoy_ids=request.buoy_ids,
        report_type=request.report_type,
        start_time=start_time,
        end_time=end_time,
        include_trends=request.include_trends
    )

    generator = ReportGenerator(config)
    pdf_bytes = generator.generate_pdf()

    # 生成文件名
    buoy_names = "_".join([
        db.query(Buoy).filter(Buoy.id == bid).first().code or bid[:8]
        for bid in request.buoy_ids
    ])
    report_type_names = {
        ReportType.DAILY: "日报",
        ReportType.WEEKLY: "周报",
        ReportType.MONTHLY: "月报",
        ReportType.QUARTERLY: "季报"
    }
    type_name = report_type_names.get(request.report_type, "报表")
    filename = f"OBMAP_{type_name}_{buoy_names}_{start_time.strftime('%Y%m%d')}_{end_time.strftime('%Y%m%d')}.pdf"

    from starlette.datastructures import URL
    import urllib.parse

    encoded_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
